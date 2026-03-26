import sys
from pathlib import Path

sys.path.append(str(Path(__file__).parent.parent / "carga"))

from openpyxl import Workbook
from carga_registro import CargarData
from Estilos.Estilos_Excel import GeneraEstilos_Excel
from Graficas.Generador_Graficas import GeneraGraficas


class GeneradorExcel:
    def __init__(self, filename: str):
        self.filename = filename
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self._sheets: dict[str, object] = {}
        self.estilos = GeneraEstilos_Excel()

    def _write_headers(self, ws, headers: list[str]) -> None:
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self.estilos.HEADER_FILL
            cell.font = self.estilos.HEADER_FONT
            cell.alignment = self.estilos.HEADER_ALIGN
            cell.border = self.estilos.BORDER

        ws.auto_filter.ref = ws.dimensions

    def _write_rows(self, ws, rows: list[list], start_row: int = 2) -> None:
        for row_idx, row_data in enumerate(rows, start_row):
            is_even_row = (row_idx - start_row) % 2 == 0
            for col_idx, value in enumerate(row_data, 1):
                # Conversión de seguridad para tipos que openpyxl no maneja (ej. pandas.Period)
                if hasattr(value, "__str__") and "Period" in str(type(value)):
                    value = str(value)
                
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.estilos.CELL_FONT
                cell.alignment = self.estilos.CELL_ALIGN
                cell.border = self.estilos.BORDER
                if is_even_row:
                    cell.fill = self.estilos.FILL_ALT

    def _autofit_columns(self, ws, headers: list[str], rows: list[list]) -> None:
        from openpyxl.utils import get_column_letter

        for col, header in enumerate(headers, 1):
            max_length = len(str(header))
            for row in rows:
                cell_value = str(row[col - 1]) if col - 1 < len(row) else ""
                max_length = max(max_length, len(cell_value))
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = max_length + 2

    def Agregar_hojas(self, title: str, headers: list[str], rows: list[list]):
        """Agrega una hoja con encabezados, datos y autoajuste de columnas."""
        idx = len(self._sheets)

        ws = self.wb.create_sheet(title)
        ws.row_dimensions[1].height = 22

        self._write_headers(ws, headers)
        self._write_rows(ws, rows, start_row=2)
        self._autofit_columns(ws, headers, rows)
        self._sheets[title] = ws
        return ws

    def save(self) -> None:
        """Guarda el archivo Excel."""
        self.wb.save(self.filename)
        print(f"Archivo guardado: {self.filename}")


# Ejemplo de uso
Cargadatos = CargarData()
Graficos = GeneraGraficas()
datos = Cargadatos.CargarDato("../CSV/fuel_prices_1970_2026.csv")
promediosanno = Cargadatos.FuncionesDataFrame(datos, "1")

Promediomes = Cargadatos.FuncionesDataFrame(datos, "2")
diferencia_anno = Cargadatos.FuncionesDataFrame(datos, "3")
variacion_diferencia = Cargadatos.FuncionesDataFrame(datos, "4")
media = Cargadatos.FuncionesDataFrame(datos, "5")
resumen_estadistico = Cargadatos.FuncionesDataFrame(datos, "6")

manager = GeneradorExcel("Generados/reporte_Costos_petroleo.xlsx")

ws_promedio = manager.Agregar_hojas(
    title="Promedio por año",
    headers=list(promediosanno.columns),
    rows=promediosanno.values.tolist(),
)

chart = Graficos.Grafico_linea(
    ws=ws_promedio,
    min_col=2,
    min_row=1,
    max_row=len(promediosanno) + 1,
    title="Precio promedio anual del petróleo",
    y_title="Precio (USD)",
    x_title="Año",
    width=24,   # Un poco más ancho para albergar los años
    height=12,
    tick_skip=5  # Salta cada 5 años para que no se amontonen las etiquetas
)
ws_promedio.add_chart(chart, "D2")

ws_mes = manager.Agregar_hojas(
    title="Promedio por mes",
    headers=list(Promediomes.columns),
    #rows=Promediomes.astype(str).values.tolist(),
    rows=Promediomes.values.tolist()
)

chart_mes = Graficos.Grafico_areas(
    ws=ws_mes,
    min_col=2,
    min_row=1,
    max_row=len(Promediomes) + 1,
    title="Precio promedio mensual del petróleo",
    y_title="Precio (USD)",
    x_title="Mes/Año",
    width=24,
    height=12,
    tick_skip=60  # Mostramos cada 5 años para que las etiquetas respiren
)
ws_mes.add_chart(chart_mes, "D2")

df_rango = diferencia_anno.reset_index()
rango = manager.Agregar_hojas(
    title="Max Min por año",
    headers=list(df_rango.columns),
    rows=df_rango.values.tolist(),
)

chart_rango = Graficos.Barras_Rango(
    ws=rango,
    min_col=2,
    min_row=1,
    max_row=len(diferencia_anno) + 1,
    title="Precio promedio mensual del petróleo",
    y_title="Precio (USD)",
    x_title="Mes/Año",
    width=24,
    height=12,
    tick_skip=5  # Mostramos cada 5 años para que las etiquetas respiren
)
rango.add_chart(chart_rango, "F2")

# Pestaña de Variaciones (Barras Divergentes)
df_variaciones = variacion_diferencia.reset_index()
diver = manager.Agregar_hojas(
    title="Variaciones",
    headers=list(df_variaciones.columns),
    rows=df_variaciones.values.tolist(),
)

chart_diver = Graficos.barras_divergente(
    ws=diver,
    min_col=3,      # Datos (Variación en Col 3)
    cat_col=2,      # Categorías (Mes/Año en Col 2)
    min_row=1,
    max_row=len(df_variaciones) + 1,
    title="Variación del Precio del Petróleo",
    y_title="Diferencia",
    x_title="Mes/Año",
    width=24,
    height=12,
    tick_skip=60 # Salto de 5 años si es mucha data
)
diver.add_chart(chart_diver, "H2")

df_decada = resumen_estadistico.reset_index()
caja = manager.Agregar_hojas(
    title="Estadisticas por decada",
    headers=list(df_decada.columns),
    rows=df_decada.values.tolist(), # Eliminado astype(str) para que el gráfico funcione
)

chart_box = Graficos.box_plot(
    ws=caja,
    min_col=2,
    max_col=len(df_decada.columns), # Toma todas las estadísticas
    min_row=1,
    max_row=len(df_decada) + 1,
    title="Estadísticas de Precios por Década",
    y_title="Precio (USD)",
    x_title="Década",
    width=24,
    height=12,
    tick_skip=1 # Pocas décadas, mostramos todas
)
caja.add_chart(chart_box, "K2")


# ws.add_chart(chart, "D2")
manager.save()
