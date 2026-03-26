from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter


class GeneraEstilos_Excel:
    def __init__(self):
        self.TAB_COLORS = ["4472C4", "ED7D31", "A9D18E", "FF0000", "7030A0"]
        self.HEADER_FILL = PatternFill("solid", fgColor="2F5496")
        self.HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        self.HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
        self.CELL_FONT = Font(name="Arial", size=10)
        self.CELL_ALIGN = Alignment(horizontal="left", vertical="center")
        self.THIN = Side(style="thin", color="BFBFBF")
        self.BORDER = Border(
            left=self.THIN, right=self.THIN, top=self.THIN, bottom=self.THIN
        )
        self.FILL_ALT = PatternFill("solid", fgColor="DCE6F1")
